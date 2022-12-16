from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill

def apply_border(ws, cell_r, cell_c):
	border_style = Side(border_style="thin", color="000000")

	# converting 0-based indexing to 1-based indexing for openpyxl
	ws.cell(
		row = 1 + cell_r,
		column = 1 + cell_c
	).border = Border(top=border_style, left=border_style, right=border_style, bottom=border_style)

def apply_highlight(ws, cell_r, cell_c, color = "FFFFFF00"):
	ws.cell(row=1 + cell_r, column=1 + cell_c).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

def compare_int_vals(ws, r1, c1, r2, c2):
	val1 = int(ws.cell(row=r1 + 1, column=c1 + 1).value)
	val2 = int(ws.cell(row=r2 + 1, column=c2 + 1).value)

	if val1 > val2:
		return 1
	elif val1 < val2:
		return -1
	else:
		return 0

def format_output_file(ws, shapes, nOctants):
	### main function

	last_col = shapes["data"][1] + 1

	# formatting data_oc + data_rank_a
	for c in range(1, shapes["data_oc_rank"][1]):
		for r in range(shapes["data_oc"][0] + 1):
			apply_border(
				ws,
				r,
				last_col + c
			)
	
	# rank 1 cell highlighting
	for c in range(shapes["data_oc_rank"][1] - shapes["data_oc"][1] - 2):
		for r in range(1, shapes["data_oc"][0] + 1):
			if(ws.cell(row=r + 1, column=last_col + shapes["data_oc"][1] + c + 1).value == 1):
				apply_highlight(ws, r, last_col + shapes["data_oc"][1] + c)

	# formatting data_rank_b
	for c in range(shapes["data_rank_b"][1]):
		for r in range(shapes["data_rank_b"][0] + 1):
			apply_border(
				ws,
				shapes["data_oc"][0] + 2 + r,
				last_col + (shapes["data_oc_rank"][1] - shapes["data_rank_b"][1] - 1) + c
			)

	last_col += shapes["data_oc_rank"][1] + 1
	
	# formatting data_tc (multiple)
	start = 0
	while(start < shapes["data_oc"][0]):
		for r in range(2, nOctants + 3):
			row = r + start*(nOctants + 3 + 2)
			greatest_col = last_col + 2

			for c in range(1, shapes["data_tc"][1]):
				col = last_col + c
				
				apply_border(ws, row, col)

				if c > 1 and r > 2 and compare_int_vals(ws, row, col, row, greatest_col) == 1:
					greatest_col = col
	
			if r > 2:
				apply_highlight(ws, row, greatest_col)
		
		start += 1
	
	last_col += shapes["data_tc"][1] + 1
	
	# formatting data_ls
	for c in range(shapes["data_ls"][1]):
		for r in range(shapes["data_ls"][0] + 1):
			apply_border(ws, r, last_col + c)
	
	last_col += shapes["data_ls"][1] + 1
	
	# formatting data_lst
	for c in range(shapes["data_lst"][1]):
		for r in range(shapes["data_lst"][0] + 1):
			apply_border(ws, r, last_col + c)